import sys
import uuid

from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

import os
import zipcodes

from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

print("edit this line if necessary, my chrome river is located at : \"C:\Program Files (x86)\"")
print("enter valid path of chrome driver in your system")
os.environ['PATH'] = r"C:\Program Files (x86)"

driver = webdriver.Chrome()

chrome_options = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
webdriver.Chrome(options=chrome_options)

# add try catch
driver.get(
    "https://www.tred.com/buy?body_style=&distance=50&exterior_color_id=&make=&miles_max=100000&miles_min=0&model=&page_size=24&price_max=100000&price_min=0&query=&requestingPage=buy&sort=desc&sort_field=updated&status=active&year_end=2022&year_start=1998&zip=")

print(driver.title)

radius = driver.find_element(By.CSS_SELECTOR, 'div.form-group.inline.radius')
select_radius = radius.find_elements(By.TAG_NAME, 'option')

radius_values = []

for radius_value in select_radius:
    value = radius_value.get_attribute('value')

    radius_values.append(value)

zip_input = driver.find_element(By.CSS_SELECTOR, 'div.form-group.inline.zip')
zip_search = zip_input.find_element(By.TAG_NAME, 'input')


user_zip_input = input("enter zip code >> ")
while not zipcodes.is_real(user_zip_input):
    print("the zipcode is invalid,  try again")
    print('>> ', end='')
    user_zip_input = input()


user_radius_input = input('enter radius >> ')
while user_radius_input not in radius_values:
    print(" invalid radius ,  try again")
    print(">>", end='')
    user_radius_input = input()


# generate filename
# file index - to unique

file_id = uuid.uuid4().time_low
filename = 'file_radius_'+user_radius_input+'zip_'+user_zip_input+'_'+str(file_id)+'.xlsx'

# prepare file

workbook = Workbook()
sheet = workbook.active

sheet['A1'] = 'Name'
sheet['B1'] = 'Price'
sheet['C1'] = 'Vehicle summary'
sheet['D1'] = 'Vehicle options'

workbook.save(filename=filename)


radius_index = radius_values.index(user_radius_input)

# fetch items
zip_search.send_keys(str(user_zip_input))
zip_search.send_keys(Keys.RETURN)

select_radius[radius_index].click()


def get_price(web_driver):
    try:
        # price
        price_box = WebDriverWait(web_driver, 90).until(
            expected_conditions.presence_of_element_located((By.CLASS_NAME, "price-box"))
        )
        price = WebDriverWait(price_box, 60).until(
            expected_conditions.presence_of_element_located((By.TAG_NAME, 'h2'))
        )

    except Exception as ex:
        print(ex)

    return price.text


def get_name(web_driver):
    try:
        # name
        car_name_elem = WebDriverWait(web_driver, 60).until(
            expected_conditions.presence_of_element_located((By.CSS_SELECTOR, 'h1.bigger.no-top-margin.hidden-xs'))
        )

        car_name = car_name_elem.text.split()
        del car_name[0]

        del car_name[-2: len(car_name)]
        car_name = ' '.join(car_name)

    except Exception as ex:
        print(ex)

    return car_name


def get_summary(web_driver):
    summary = []

    try:
        # summary table
        summary_table = WebDriverWait(web_driver, 60).until(
            expected_conditions.presence_of_all_elements_located((By.ID, "summary-table"))
        )

        # summary_table -get all table rows for this table
        summary_table_rows = WebDriverWait(summary_table[1], 60).until(
            expected_conditions.presence_of_all_elements_located((By.TAG_NAME, 'tr'))
        )

        # fetch content in each table row

        for summary_table_row in range(1, len(summary_table_rows)):
            t_head = WebDriverWait(summary_table_rows[summary_table_row], 150).until(
                expected_conditions.presence_of_element_located((By.TAG_NAME, 'th'))
            )

            t_data = WebDriverWait(summary_table_rows[summary_table_row], 150).until(
                expected_conditions.presence_of_element_located((By.TAG_NAME, 'td'))
            )

            summary.append(t_head.text + t_data.text)
            # print("{} {}".format(t_head.text, t_data.text), end="\n")

    except Exception as ex:
        print(ex)

    summary = ' '.join(summary)
    summary = '['+summary + ']'
    return summary


def get_options(web_driver):
    options = []
    try:
        # there are two 'options-table' , this line return the first one
        options_table = WebDriverWait(web_driver, 150).until(
            expected_conditions.presence_of_element_located((By.ID, 'options-table'))
        )

        options_row = WebDriverWait(options_table, 150).until(
            expected_conditions.presence_of_all_elements_located((By.TAG_NAME, 'tr'))
        )

        index = 0

        for option in options_row:
            if 'Options' in option.text:
                index = index + 1
                break

            else:
                index = index + 1

        for option in range(index, len(options_row)):
            options.append(options_row[option].text)
            print(options_row[option].text, end='\n')

    except Exception as ex:
        print(ex)

    options = ' '.join(options)
    options = '['+options + ']'
    return options


def write_to_file(vehicle_name, vehicle_price, vehicle_summary, vehicle_options):
    car_workbook = load_workbook(filename=filename)
    active_sheet = car_workbook.active
    # row_data = [vehicle_name, vehicle_price, vehicle_summary, vehicle_options]

    # get the current row
    current_row = active_sheet.max_row + 1

    active_sheet.cell(row=current_row, column=1).value = vehicle_name
    active_sheet.cell(row=current_row, column=2).value = vehicle_price
    active_sheet.cell(row=current_row, column=3).value = vehicle_summary
    active_sheet.cell(row=current_row, column=4).value = vehicle_options

    car_workbook.save(filename=filename)


# fetch all the car_divs which contains cars to sell
# some of the containers are for adverts so we have to skip them
try:
    car_divs = WebDriverWait(driver, 60).until(
        expected_conditions.presence_of_all_elements_located((By.XPATH, '//*[contains(@href,"/buy/")]'))
    )
except TimeoutException as time_exception:
    print("you have this exception because there were no available matches for zip code: {} and radius: {}".format(user_zip_input, user_radius_input))
    print("run the script again with different values")
    sys.exit()

# iterate through all the car_divs and click one by one to fetch the required details
for car_div in range(0, len(car_divs)):
    # reload the car links
    try:
        car_link = WebDriverWait(driver, 120).until(
            expected_conditions.presence_of_all_elements_located((By.XPATH, '//*[contains(@href,"/buy/")]'))
        )
    except TimeoutException as ex:
        print("timeout exception")
        sys.exit()

    car_link[car_div].click()

    name = get_name(driver)
    car_price = get_price(driver)
    car_summary = get_summary(driver)
    options_data = get_options(driver)

    write_to_file(name, car_price, car_summary, options_data)
    print("name: {}, price: {}, summary: {}, options: {}".format(name, car_price, car_summary, options_data), end='\n')

    # navigate back the previous page which displayed all the cars according to user input (zip_code and radius)
    driver.back()
    car_div = car_div + 1


print("result filename is {} in the current directory".format(filename))
